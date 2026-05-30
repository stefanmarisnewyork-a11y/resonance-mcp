"""ResonanceOS MCP — single-file deployment build (flat, no package folders)."""
from __future__ import annotations
import os, re, json, time, sqlite3, hashlib, math, csv, io
from pathlib import Path
from abc import ABC, abstractmethod
from mcp.server.fastmcp import FastMCP
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ===================== from processing.py =====================
"""RMT DB processing logic — the computation that lives INSIDE the database layer.

Per the ResonanceOS canon: the RMT DB owns all logic (scoring, resonance
translation, alignment, universe/dedup math) but operates only on its fixed
input/output spec. The Resonance Engine does NOT compute — it only conforms
input to this spec and formats output.

The dummy -> real R swap happens HERE, inside the DB's processor, not in a
separate engine. DummyProcessor today; RealRProcessor (R on AWS) at Stage 4.
"""


MOTIVATIONS = [
    "Transcendence", "Self-Knowledge", "Creativity", "Love", "Power",
    "Good Life", "Heroism", "Wealth", "Status", "Fitness", "Competency",
    "Aspiration", "Achievement", "Belonging", "Security", "Self-Esteem",
]
FLOOR_THRESHOLD = 20


class Processor(ABC):
    """The DB's processing contract. Real R must satisfy exactly this."""

    @abstractmethod
    def score_creative_ref(self, ref: str) -> dict: ...
    @abstractmethod
    def score_asset(self, asset_id: str, kind: str) -> dict: ...

    # shared (spec-level) math — identical across dummy/real, operates on conformed data
    def resonance(self, a: dict, b: dict) -> float:
        cv = [a.get(m, 0.0) for m in MOTIVATIONS]
        pv = [b.get(m, 0.0) for m in MOTIVATIONS]
        dot = sum(x * y for x, y in zip(cv, pv))
        nc = math.sqrt(sum(x * x for x in cv)) or 1e-9
        npv = math.sqrt(sum(y * y for y in pv)) or 1e-9
        return round(dot / (nc * npv) * 100, 1)

    def alignment(self, creative: dict, audience: dict, threshold: float = 0.20) -> dict:
        per, n = {}, 0
        for m in MOTIVATIONS:
            ok = abs(creative.get(m, 0) - audience.get(m, 0)) <= threshold
            per[m] = ok; n += int(ok)
        return {"aligned_count": n, "total": len(MOTIVATIONS),
                "match_score": round(self.resonance(creative, audience) / 100, 3),
                "per_spoke": per}

    def activate(self, scores_pct: dict, universes: dict, threshold: int = FLOOR_THRESHOLD,
                 overlap: float = 0.62, cap: int | None = None) -> dict:
        t = max(int(threshold), FLOOR_THRESHOLD)
        active = [m for m in scores_pct if scores_pct[m] >= t]
        naive = sum(universes.get(m, 0) for m in active)
        dedup = int(naive * overlap)
        if cap is not None:
            dedup = min(dedup, cap)        # unique reach can never exceed the audience total
        return {"threshold": t, "active": active, "active_count": len(active),
                "naive_universe": naive, "dedup_universe": dedup}


class DummyProcessor(Processor):
    """Stand-in math. Stable, seeded off ids/refs. Swap for RealRProcessor at Stage 4."""

    def _seed(self, *keys, spiky=1.6) -> dict:
        out = {}
        for m in MOTIVATIONS:
            h = hashlib.sha256(("|".join(keys + (m,))).encode()).digest()
            out[m] = round((int.from_bytes(h[:4], "big") / 0xFFFFFFFF) ** spiky, 3)
        return out

    def score_creative_ref(self, ref: str) -> dict:
        return self._seed(ref, "creative")

    def score_asset(self, asset_id: str, kind: str) -> dict:
        return self._seed(asset_id, kind, spiky=1.6 if kind == "creative" else 0.9)


# ===================== from specs.py =====================
"""Declared input specs — the single source of truth for what the DB accepts.

Today: the ad-creative spec. Structured so audience / program-profile / campaign
specs can follow the same shape later. The server exposes this (get_input_spec),
validates against it (validate_creative), and Claude reads it to map messy input
toward a known contract. One definition, three consumers.
"""


CREATIVE_INPUT_SPEC = {
    "spec": "ad_creative",
    "version": "1.0",
    "description": "An ad creative scored as a motivational profile: a percentage "
                   "(0-100) on each of zero-or-more of the 16 motivations.",
    "fields": {
        "creative": {
            "type": "string", "required": True,
            "description": "Creative name or id. Matched to a queued creative if it "
                           "exists, otherwise a new creative is created.",
        },
        "scores": {
            "type": "object", "required": True,
            "key_domain": MOTIVATIONS,
            "value_type": "number", "value_range": [0, 100],
            "description": "Map of motivation -> percentage. Only the 16 listed "
                           "motivations are valid keys. Omitted motivations = 0. "
                           "A creative may be active on 1 to 16 of them.",
        },
    },
    "motivations": MOTIVATIONS,
    "rules": [
        "Every key in 'scores' must be exactly one of the 16 motivations (case-sensitive).",
        "Every value must be a number between 0 and 100 inclusive.",
        "Unknown motivations or out-of-range values cause the whole batch to be rejected.",
        "'creative' must be non-empty.",
    ],
}


def validate_creative(item: dict) -> list[str]:
    """Validate one {creative, scores} item against the spec. Returns a list of problems
    (empty = valid). This is THE validation; tools call it instead of ad-hoc checks."""
    problems = []
    label = str(item.get("creative", "")).strip()
    if not label:
        problems.append("missing 'creative' (name or id)")
    raw = item.get("scores", {})
    if not isinstance(raw, dict):
        problems.append(f"{label or '?'}: 'scores' must be an object {{motivation: 0-100}}")
        return problems
    unknown = sorted(set(raw) - set(MOTIVATIONS))
    if unknown:
        problems.append(f"{label}: unknown motivations {unknown} — must be one of the 16")
    oor = {k: v for k, v in raw.items()
           if not (isinstance(v, (int, float)) and 0 <= v <= 100)}
    if oor:
        problems.append(f"{label}: out-of-range values {oor} (must be 0-100)")
    return problems


# ===================== from translate.py =====================
"""Resonance Engine layer — the translation boundary (ENG).

Everything here turns messy outside-world input into DB-conforming data, or
renders DB data back out. The clean core (db.py) never sees anything that hasn't
passed through here. This is the literal implementation of the tagging rule:
  ENG = translate / conform / map / reject / render
  DB  = operate on already-conforming data
"""



# ---- filename metadata mining (the "add back metadata" capability) ----
def mine_filename(filename: str) -> dict:
    """Extract structured metadata from a dense creative filename.

    Real-world example:
      JRCL2669000H_Fight_Value_CJR_GM_599_HBCWesternBacon_HD15S_INNOVID.mp4
    Conventions drift file-to-file, so this is heuristic, not a rigid parser —
    in production the LLM (Claude) does this read and flags anything ambiguous.
    Everything uncertain goes into 'flags' rather than being silently guessed.
    """
    stem = re.sub(r"\.(mp4|mov|webm|xml)$", "", filename, flags=re.I)
    parts = stem.split("_")
    meta: dict = {"raw_filename": filename}
    flags: list[str] = []

    brand_map = {"CJR": "Carl's Jr.", "HBC": "Hardee's"}
    for p in parts:
        if p in brand_map:
            meta["brand"] = brand_map[p]
        if re.fullmatch(r"\d{3}", p):                 # 599 -> $5.99
            meta["price"] = f"${int(p)/100:.2f}"
        m = re.fullmatch(r"(?:HD)?(\d+)S", p, flags=re.I)  # HD15S -> 15s
        if m:
            meta["duration_s"] = int(m.group(1))
        if p.upper() in {"INNOVID", "FREEWHEEL", "SAMBA"}:
            meta["ad_server"] = p.upper().title()
    if re.match(r"^[A-Z]{2,}\d{4,}", parts[0]):
        meta["creative_code"] = parts[0]

    # human-readable product guess = the longest CamelCase-ish token
    cand = [p for p in parts if re.search(r"[A-Z][a-z]", p)]
    if cand:
        meta["product"] = re.sub(r"(?<!^)(?=[A-Z])", " ", max(cand, key=len)).strip()

    for need in ("brand", "price", "duration_s"):
        if need not in meta:
            flags.append(f"could not mine '{need}' from filename — NEEDS REVIEW")
    meta["flags"] = flags
    return meta


# ---- creative reference classification (translation/boundary) ----
def classify_creative_ref(ref: str) -> tuple[str, str]:
    """Classify a raw-creative reference. Returns (kind, normalized_ref).

    kind: 'vast' | 'file' | 'url' | 'unknown'. This is boundary work — deciding
    what the outside-world string actually is before the engine tries to score it.
    """
    r = ref.strip()
    low = r.lower()
    if low.startswith(("http://", "https://")):
        if "vast" in low or low.endswith(".xml") or "/vast" in low:
            return "vast", r
        return "url", r
    if low.endswith((".mp4", ".mov", ".webm", ".xml")) or "/" in r or "\\" in r:
        return "file", r
    return "unknown", r


# ---- 1.1.2 multi-database identification ----
DB_SIGNATURES = {
    "US_CONTENT": ["us", "content"],
    "CANADA_CONTENT": ["canada", "ca_content", "cancontent"],
    "US_MEDIAPROBE": ["mediaprobe", "mp_id", "mprobe"],
    "US_LUMEN": ["lumen", "fixation", "eyetrack"],
}


def identify_database(filename: str, header: str) -> tuple[str, str]:
    """Sniff format/header to route a file. Returns (db_name, reason)."""
    hay = f"{filename} {header}".lower()
    for db_name, sigs in DB_SIGNATURES.items():
        for s in sigs:
            if s in hay:
                return db_name, f"matched signature '{s}'"
    return "UNKNOWN", "no signature matched — flag for review"


# ---- 1.1.1 ingest & translation ----
ISCII_RE = re.compile(r"[^\x09\x0A\x0D\x20-\x7E]")  # keep tab/newline/printable ASCII


def conform_row(row: dict, known_genre: str | None = None) -> dict:
    """Take one messy row, return a conforming row + any flags.

    Deterministic cleaning (ISCII strip, trim) is pure code. Genuinely ambiguous
    cases are surfaced as `flags` for the agent (Claude) to reason about — that is
    the judgment half of the split.
    """
    cleaned, flags = {}, []
    for k, v in row.items():
        key = ISCII_RE.sub("", str(k)).strip()
        val = ISCII_RE.sub("", str(v)).strip() if v is not None else ""
        cleaned[key] = val

    # blank value-signals with a known genre -> eligible for genre-average fallback
    blank_vs = [k for k, v in cleaned.items() if k.lower().startswith("vs") and v == ""]
    if blank_vs:
        if known_genre:
            flags.append(
                f"{len(blank_vs)} blank value-signals; genre='{known_genre}' "
                f"-> apply PVS Norms genre-average fallback in DB"
            )
        else:
            flags.append(
                f"{len(blank_vs)} blank value-signals and no genre -> NEEDS REVIEW"
            )
    return {"row": cleaned, "flags": flags}


# ---- 1.1.10 resonance-match polygon (SVG) ----
def polygon_svg(audience: dict[str, float],
                creatives: list[tuple[str, dict[str, float], dict]]) -> str:
    """Render the 16-spoke radar: shaded audience profile + creative overlays.

    creatives: list of (label, scores, alignment_result).
    Returns a standalone SVG string the agent renders inline.
    """
    cx, cy, R = 260, 260, 200
    n = len(MOTIVATIONS)
    palette = ["#1D9E75", "#BA7517", "#185FA5", "#993556"]

    def pt(i, r):
        ang = -math.pi / 2 + 2 * math.pi * i / n
        return cx + r * math.cos(ang) * 1.0, cy + r * math.sin(ang)

    parts = [f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 520 560">']
    # grid rings
    for ring in (0.25, 0.5, 0.75, 1.0):
        pts = " ".join(f"{x:.1f},{y:.1f}" for x, y in (pt(i, R * ring) for i in range(n)))
        parts.append(f'<polygon points="{pts}" fill="none" stroke="#D3D1C7" stroke-width="0.5"/>')
    # spokes + labels
    for i, m in enumerate(MOTIVATIONS):
        x, y = pt(i, R)
        lx, ly = pt(i, R + 22)
        parts.append(f'<line x1="{cx}" y1="{cy}" x2="{x:.1f}" y2="{y:.1f}" stroke="#E8E6DD" stroke-width="0.5"/>')
        anchor = "middle" if abs(lx - cx) < 30 else ("start" if lx > cx else "end")
        parts.append(f'<text x="{lx:.1f}" y="{ly:.1f}" font-size="9" fill="#5F5E5A" '
                     f'text-anchor="{anchor}" font-family="Arial">{m}</text>')
    # audience filled polygon
    apts = " ".join(f"{x:.1f},{y:.1f}" for x, y in (pt(i, R * audience.get(m, 0)) for i, m in enumerate(MOTIVATIONS)))
    parts.append(f'<polygon points="{apts}" fill="#7F77DD" fill-opacity="0.28" stroke="#534AB7" stroke-width="1"/>')
    # creative overlays (dashed)
    for j, (label, scores, _al) in enumerate(creatives):
        col = palette[j % len(palette)]
        cpts = " ".join(f"{x:.1f},{y:.1f}" for x, y in (pt(i, R * scores.get(m, 0)) for i, m in enumerate(MOTIVATIONS)))
        parts.append(f'<polygon points="{cpts}" fill="none" stroke="{col}" stroke-width="1.5" stroke-dasharray="5 3"/>')
    # legend
    ly = 520
    parts.append(f'<rect x="60" y="{ly-8}" width="12" height="12" fill="#7F77DD" fill-opacity="0.28" stroke="#534AB7"/>')
    parts.append(f'<text x="78" y="{ly+2}" font-size="10" fill="#2C2C2A" font-family="Arial">Audience profile</text>')
    parts.append('</svg>')
    return "".join(parts)


# ===================== from reports.py =====================
"""Standard report rollups — registry feature 1.1.3.

STAGE 1 (honest architecture): program motivational profiles are READ from the
RMT DB (program_profiles table); resonance is COMPUTED by matching the creative's
profile against each program's stored profile. Real shape:
  DB stores profiles (truth) -> engine computes translation -> report.
Matching math is a stand-in (cosine similarity, 0-100). Stage 4 swaps in real R.
Network and network+daypart reports are aggregations over per-program results.
"""



def _resonance(creative_scores, program_scores) -> float:
    cv = [creative_scores.get(m, 0.0) for m in MOTIVATIONS]
    pv = [program_scores.get(m, 0.0) for m in MOTIVATIONS]
    dot = sum(a * b for a, b in zip(cv, pv))
    nc = math.sqrt(sum(a * a for a in cv)) or 1e-9
    npv = math.sqrt(sum(b * b for b in pv)) or 1e-9
    return round(dot / (nc * npv) * 100, 1)


def program_report(creative_scores, programs):
    headers = ["Program", "Network", "Daypart", "Resonance Index", "Reach (000s)"]
    rows = [[p["program"], p["network"], p["daypart"],
             _resonance(creative_scores, p["scores"]), p["reach"]] for p in programs]
    rows.sort(key=lambda r: r[3], reverse=True)
    return headers, rows


def network_report(creative_scores, programs):
    agg = {}
    for p in programs:
        r = _resonance(creative_scores, p["scores"])
        agg.setdefault(p["network"], [[], 0])
        agg[p["network"]][0].append(r)
        agg[p["network"]][1] += p["reach"]
    headers = ["Network", "Avg Resonance Index", "Total Reach (000s)", "Programs"]
    rows = [[net, round(sum(rs) / len(rs), 1), reach, len(rs)]
            for net, (rs, reach) in agg.items()]
    rows.sort(key=lambda r: r[1], reverse=True)
    return headers, rows


def network_daypart_report(creative_scores, programs):
    agg = {}
    for p in programs:
        r = _resonance(creative_scores, p["scores"])
        key = (p["network"], p["daypart"])
        agg.setdefault(key, [[], 0])
        agg[key][0].append(r)
        agg[key][1] += p["reach"]
    headers = ["Network", "Daypart", "Avg Resonance Index", "Total Reach (000s)"]
    rows = [[net, dp, round(sum(rs) / len(rs), 1), reach]
            for (net, dp), (rs, reach) in agg.items()]
    rows.sort(key=lambda r: r[2], reverse=True)
    return headers, rows


REPORT_TYPES = {
    "programs": "AD_MOT \u2192 Programs",
    "networks": "AD_MOT \u2192 Networks",
    "networks_daypart": "AD_MOT \u2192 Networks + Daypart",
}


# ===================== from excel_report.py =====================
"""Excel report writer — three-tab AD_MOT workbook computed against stored profiles."""


HEADER_BG = "26215C"
ALT = "F4F3FA"
OUT_DIR = Path("/mnt/user-data/outputs")


def build_workbook(creative_id, creative_name, creative_scores, programs, metadata=None):
    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="D3D1C7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    cov = wb.create_sheet("Summary")
    cov["A1"] = "ResonanceOS \u2014 AD_MOT Standard Reports"
    cov["A1"].font = Font(name="Arial", bold=True, size=14, color=HEADER_BG)
    info = [("Creative", creative_name), ("Creative ID", creative_id),
            ("Programs in DB", str(len(programs)))]
    if metadata:
        for k in ("brand", "product", "price", "duration_s", "ad_server"):
            if metadata.get(k):
                lbl = {"duration_s": "Duration (s)", "ad_server": "Ad Server"}.get(k, k.title())
                info.append((lbl, str(metadata[k])))
    r = 3
    for label, val in info:
        cov.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
        cov.cell(row=r, column=2, value=val).font = Font(name="Arial", size=10)
        r += 1
    cov.cell(row=r + 1, column=1,
             value="Resonance computed vs. stored program profiles (RMT DB).").font = Font(
        name="Arial", italic=True, size=9, color="5F5E5A")
    cov.column_dimensions["A"].width = 16
    cov.column_dimensions["B"].width = 42

    builders = {
        "By Program": R.program_report,
        "By Network": R.network_report,
        "By Network + Daypart": R.network_daypart_report,
    }
    for tab, fn in builders.items():
        headers, rows = fn(creative_scores, programs)
        ws = wb.create_sheet(tab)
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor=HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for i, row in enumerate(rows, start=2):
            ws.append(row)
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=i, column=c)
                cell.font = Font(name="Arial", size=10)
                cell.border = border
                if i % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=ALT)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{chr(64+len(headers))}{len(rows)+1}"
        for idx in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + idx)].width = 18

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    safe = "".join(ch for ch in creative_name if ch.isalnum() or ch in " -_").strip().replace(" ", "_")
    path = OUT_DIR / f"AD_MOT_Report_{safe or creative_id}.xlsx"
    wb.save(path)
    return str(path)


# ===================== from db.py =====================
"""RMT DB — the clean core AND the processing owner.

Per ResonanceOS canon: the DB owns all logic (scoring, resonance translation,
alignment, universe math) but operates only on its fixed spec. The Resonance
Engine conforms input to this spec and formats output — it does not compute.
The dummy->real R swap lives inside self.processor, INSIDE this layer.
SQLite today; real RMT DB later, same interface.
"""



DB_PATH = Path(os.environ.get("DB_DIR", ".")) / "resonance.db"


class DB:
    def __init__(self, path: Path = DB_PATH, processor=None):
        self.conn = sqlite3.connect(path, check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        # the processing engine lives INSIDE the DB layer. swap to RealRProcessor at Stage 4.
        self.processor = processor or DummyProcessor()
        self._init_schema()

    def _init_schema(self) -> None:
        self.conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS creative_scores (
                asset_id TEXT PRIMARY KEY,
                name TEXT,
                scores_json TEXT NOT NULL,
                source_db TEXT,
                loaded_at REAL
            );
            CREATE TABLE IF NOT EXISTS audience_scores (
                audience_id TEXT PRIMARY KEY,
                name TEXT,
                scores_json TEXT NOT NULL,
                universe_size INTEGER,
                loaded_at REAL
            );
            CREATE TABLE IF NOT EXISTS content (
                content_id TEXT PRIMARY KEY,
                name TEXT,
                section TEXT,          -- '3A' web | '3B' tv
                scores_json TEXT,
                loaded_at REAL
            );
            CREATE TABLE IF NOT EXISTS scoring_queue (
                queue_id TEXT PRIMARY KEY,
                creative_id TEXT NOT NULL,
                name TEXT,
                metadata_json TEXT,     -- mined from filename (brand, price, product, duration...)
                video_url TEXT,
                status TEXT NOT NULL,    -- 'pending' | 'scored'
                requested_by TEXT,
                scored_by TEXT,
                scored_at REAL,
                created_at REAL
            );
            CREATE TABLE IF NOT EXISTS audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ts REAL NOT NULL,
                operation TEXT NOT NULL,
                detail TEXT
            );
            CREATE TABLE IF NOT EXISTS program_profiles (
                program TEXT PRIMARY KEY,
                network TEXT,
                daypart TEXT,
                scores_json TEXT NOT NULL,   -- pre-scored 16-motivation profile (0-1)
                reach INTEGER,
                updated_at REAL
            );
            CREATE TABLE IF NOT EXISTS campaigns (
                name TEXT PRIMARY KEY,
                created_at REAL
            );
            CREATE TABLE IF NOT EXISTS audience_universes (
                audience_id TEXT NOT NULL,
                platform TEXT NOT NULL,
                motivation TEXT NOT NULL,
                size INTEGER NOT NULL,
                as_of TEXT,            -- refresh date label, e.g. '2026-03'
                loaded_at REAL,
                PRIMARY KEY (audience_id, platform, motivation)
            );
            CREATE TABLE IF NOT EXISTS campaign_pairs (
                campaign TEXT NOT NULL,
                creative_id TEXT NOT NULL,
                audience_id TEXT NOT NULL,
                added_at REAL,
                PRIMARY KEY (campaign, creative_id, audience_id)
            );
            """
        )
        self.conn.commit()

    # ---- clean-core writes (1.1.1b) ----
    def write_creative(self, asset_id, name, scores, source_db):
        self.conn.execute(
            "INSERT OR REPLACE INTO creative_scores VALUES (?,?,?,?,?)",
            (asset_id, name, json.dumps(scores), source_db, time.time()),
        )
        self.conn.commit()

    def write_audience(self, audience_id, name, scores, universe_size):
        self.conn.execute(
            "INSERT OR REPLACE INTO audience_scores VALUES (?,?,?,?,?)",
            (audience_id, name, json.dumps(scores), universe_size, time.time()),
        )
        self.conn.commit()

    # ---- reads ----
    def get_creative(self, asset_id):
        r = self.conn.execute(
            "SELECT * FROM creative_scores WHERE asset_id=?", (asset_id,)
        ).fetchone()
        return dict(r) if r else None

    def get_audience(self, audience_id):
        r = self.conn.execute(
            "SELECT * FROM audience_scores WHERE audience_id=?", (audience_id,)
        ).fetchone()
        return dict(r) if r else None

    def query(self, sql, params=()):
        return [dict(r) for r in self.conn.execute(sql, params).fetchall()]

    # ---- audit (1.1.6) ----
    def log(self, operation: str, detail: str = "") -> None:
        self.conn.execute(
            "INSERT INTO audit_log (ts, operation, detail) VALUES (?,?,?)",
            (time.time(), operation, detail),
        )
        self.conn.commit()

    # ---- scoring queue ----
    def enqueue(self, queue_id, creative_id, name, metadata, video_url, requested_by="unknown"):
        self.conn.execute(
            "INSERT OR REPLACE INTO scoring_queue VALUES (?,?,?,?,?,?,?,?,?,?)",
            (queue_id, creative_id, name, json.dumps(metadata), video_url,
             "pending", requested_by, None, None, time.time()),
        )
        self.conn.commit()

    def get_queue_item(self, queue_id):
        r = self.conn.execute(
            "SELECT * FROM scoring_queue WHERE queue_id=?", (queue_id,)
        ).fetchone()
        return dict(r) if r else None

    def pending_queue(self):
        return self.query(
            "SELECT queue_id, creative_id, name, metadata_json, video_url, requested_by "
            "FROM scoring_queue WHERE status='pending' ORDER BY created_at"
        )

    def mark_scored(self, queue_id, scored_by):
        self.conn.execute(
            "UPDATE scoring_queue SET status='scored', scored_by=?, scored_at=? WHERE queue_id=?",
            (scored_by, time.time(), queue_id),
        )
        self.conn.commit()

    # ---- program profiles (pre-scored RMT reference data) ----
    def upsert_program(self, program, network, daypart, scores, reach):
        self.conn.execute(
            "INSERT OR REPLACE INTO program_profiles VALUES (?,?,?,?,?,?)",
            (program, network, daypart, json.dumps(scores), reach, time.time()),
        )
        self.conn.commit()

    def all_programs(self):
        return [
            {**r, "scores": json.loads(r["scores_json"])}
            for r in self.query("SELECT * FROM program_profiles ORDER BY program")
        ]

    # ============================================================
    # PROCESSING (DB owns the logic; delegates math to self.processor)
    # ============================================================
    def score_and_store_creative(self, asset_id, name, ref=None, source_db="US_CONTENT"):
        """Score a creative (dummy/real processor) and store it. ref=raw creative reference."""
        scores = (self.processor.score_creative_ref(ref) if ref
                  else self.processor.score_asset(asset_id, "creative"))
        self.write_creative(asset_id, name, scores, source_db)
        return scores

    def score_and_store_audience(self, audience_id, name, universe_size):
        scores = self.processor.score_asset(audience_id, "audience")
        self.write_audience(audience_id, name, scores, universe_size)
        return scores

    def store_graded_creative(self, asset_id, name, scores, source_db="US_CONTENT"):
        """Store an externally-graded creative profile (e.g. from Skip's scoring card)."""
        self.write_creative(asset_id, name, scores, source_db)
        return scores

    def match(self, audience_id, creative_ids):
        """Resonance match: alignment of each creative vs the audience profile."""
        aud = self.get_audience(audience_id)
        if not aud:
            return None
        aud_scores = json.loads(aud["scores_json"])
        results = []
        for cid in creative_ids:
            c = self.get_creative(cid)
            if not c:
                results.append({"creative_id": cid, "error": "not found"}); continue
            cs = json.loads(c["scores_json"])
            al = self.processor.alignment(cs, aud_scores)
            results.append({"creative_id": cid, "name": c["name"], "scores": cs,
                            "aligned": f'{al["aligned_count"]} of {al["total"]}',
                            "match_score": al["match_score"]})
        ranked = sorted([r for r in results if "match_score" in r],
                        key=lambda r: r["match_score"], reverse=True)
        if ranked:
            ranked[0]["best_match"] = True
        return {"audience": aud["name"], "audience_scores": aud_scores, "results": results}

    def program_resonance(self, creative_scores):
        """Compute creative-vs-stored-program-profile resonance for all programs."""
        return [{**p, "resonance": self.processor.resonance(creative_scores, p["scores"])}
                for p in self.all_programs()]

    # ---- campaigns ----
    def upsert_campaign(self, name):
        self.conn.execute("INSERT OR IGNORE INTO campaigns VALUES (?,?)", (name, time.time()))
        self.conn.commit()

    # ---- audience universes (monthly-refreshed reference data, Flow A) ----
    def upsert_audience_universe(self, audience_id, platform, motivation, size, as_of):
        self.conn.execute(
            "INSERT OR REPLACE INTO audience_universes VALUES (?,?,?,?,?,?)",
            (audience_id, platform, motivation, int(size), as_of, time.time()))
        self.conn.commit()

    def get_audience_universes(self, audience_id, platform):
        """Stored per-motivation sizes for an audience on a platform, + the as_of date.
        Returns ({motivation: size}, as_of) or ({}, None) if not loaded."""
        rows = self.query(
            "SELECT motivation, size, as_of FROM audience_universes "
            "WHERE audience_id=? AND platform=?", (audience_id, platform))
        sizes = {r["motivation"]: r["size"] for r in rows}
        as_of = rows[0]["as_of"] if rows else None
        return sizes, as_of

    def loaded_platforms(self, audience_id):
        return [r["platform"] for r in self.query(
            "SELECT DISTINCT platform FROM audience_universes WHERE audience_id=?", (audience_id,))]

    def add_pair(self, campaign, creative_id, audience_id):
        self.upsert_campaign(campaign)
        self.conn.execute("INSERT OR IGNORE INTO campaign_pairs VALUES (?,?,?,?)",
                          (campaign, creative_id, audience_id, time.time()))
        self.conn.commit()

    def campaign_view(self, name, platform, fallback_fn=None):
        """User-facing campaign read in the per-creative / per-motivation shape.

        Reads stored per-motivation audience sizes for the given platform from the
        audience_universes table (monthly-refreshed reference data). If a pair's
        audience has no loaded data for that platform and fallback_fn is given, uses
        it (clearly flagged as not-real). Returns as_of dates so staleness is visible.
        """
        pairs = self.query(
            "SELECT creative_id, audience_id FROM campaign_pairs WHERE campaign=? ORDER BY added_at",
            (name,))
        if not pairs:
            return None
        rows = []
        for p in pairs:
            c = self.get_creative(p["creative_id"])
            a = self.get_audience(p["audience_id"])
            if not c or not a:
                rows.append({"creative_id": p["creative_id"], "error": "creative or audience not found"})
                continue
            cs = json.loads(c["scores_json"])
            active = {m: round(cs[m] * 100) for m in cs if cs[m] > 0}
            uni, as_of = self.get_audience_universes(p["audience_id"], platform)
            data_source = "loaded"
            if not uni:
                if fallback_fn is None:
                    rows.append({"creative_id": p["creative_id"], "creative": c["name"],
                                 "audience": a["name"],
                                 "error": f"no audience-size data loaded for {a['name']} on {platform} "
                                          f"— run ingest_audience_universes"})
                    continue
                uni = fallback_fn(p["audience_id"], a["universe_size"])
                data_source = "PLACEHOLDER (not loaded — illustrative only)"
                as_of = None
            per_motiv = [{"motivation": m, "score_pct": active[m], "audience_size": uni.get(m, 0)}
                         for m in sorted(active, key=lambda m: uni.get(m, 0), reverse=True)]
            naive = sum(uni.get(m, 0) for m in active)
            act = self.processor.activate(active, uni, threshold=20, cap=a["universe_size"])
            rows.append({
                "creative_id": p["creative_id"], "creative": c["name"],
                "audience_id": p["audience_id"], "audience": a["name"],
                "motivational_profile": list(active.keys()),
                "per_motivation": per_motiv,
                "audience_size": naive,
                "dedup_audience_size": act["dedup_universe"],
                "creative_scores": cs,
                "as_of": as_of, "data_source": data_source,
            })
        return {"campaign": name, "platform": platform, "creatives": rows}


# ===================== from server.py =====================
"""ResonanceOS internal MCP server — scaffold (registry 1.1.0).

Exposes the hero-path tools over stdio for Claude Code / Cowork:
  identify_database (1.1.2) -> ingest (1.1.1) -> write_creative (1.1.1b)
  -> score_match (1.1.9 + 1.1.10) -> with audit_log (1.1.6) throughout.

ADAPTER SWAP POINTS (dummy -> real) are marked below.
"""




mcp = FastMCP("resonanceos-internal")

# The RMT DB owns processing; the dummy->real R swap lives inside its processor.
db = DB()


# ============================================================
# TOOL TIERING (for v1.2 external exposure)
# Principle: EXTERNAL callers get answers SCOPED TO THEIR OWN INPUTS (their creative,
# their campaign). INTERNAL callers additionally get tools that ENUMERATE/BROWSE the
# reference data (program library, audience universes, raw SQL) — the proprietary asset.
# Enumeration of the library is the data-leakage risk; keep it internal.
# This registry is the v1.2 allowlist: external connectors expose only EXTERNAL_SAFE.
# ============================================================
TOOL_TIER = {
    # external-safe: scoped to the caller's own creative / campaign
    "get_input_spec": "external_safe",
    "score_creative": "external_safe",
    "score_match": "external_safe",
    "preview_scored_creatives": "external_safe",
    "commit_scored_creatives": "external_safe",
    "enqueue_for_scoring": "external_safe",
    "view_scoring_queue": "external_safe",
    "open_scoring_card": "external_safe",
    "submit_scores": "external_safe",
    "add_to_campaign": "external_safe",
    "show_campaign": "external_safe",        # NOTE: external version must aggregate (see below)
    "generate_report": "external_safe",      # NOTE: external version = network/daypart only
    "setup_instructions": "external_safe",
    "write_audience": "external_safe",
    "write_creative": "external_safe",
    # internal-only: enumerate / browse the proprietary reference data
    "programs_by_motivation": "internal_only",
    "view_program_profiles": "internal_only",
    "ingest_program_profiles": "internal_only",
    "ingest_audience_universes": "internal_only",
    "query_db": "internal_only",             # raw SQL — never external
    "audit_tail": "internal_only",
}

# v1.2 product decisions recorded now (not enforced until external exposure exists):
EXTERNAL_GRANULARITY = {
    "show_campaign": "aggregate to network/daypart for external; named-show resonance is internal-only",
    "generate_report": "external AD_MOT report = network + daypart tabs ONLY; drop the named-Program tab",
    "rationale": "named-show-level scores are the proprietary content asset; protect from third parties",
}


@mcp.tool()
def list_tools_by_tier() -> str:
    """[INTERNAL] Show how each tool is classified for external (v1.2) exposure:
    external_safe (scoped to caller's own inputs) vs internal_only (enumerates the
    proprietary reference data). This is the v1.2 allowlist + recorded granularity rules."""
    ext = sorted(t for t, v in TOOL_TIER.items() if v == "external_safe")
    intl = sorted(t for t, v in TOOL_TIER.items() if v == "internal_only")
    return json.dumps({"external_safe": ext, "internal_only": intl,
                       "external_granularity_rules": EXTERNAL_GRANULARITY,
                       "note": "Tiering is declared now; enforcement (filtering the external "
                               "tool surface) lands with v1.2 auth/multi-tenancy."})

# Platform activation haircuts — reachable audience depends on the DSP/SSP selected.
# PLACEHOLDER factors; real numbers come from each platform later (e.g. FreeWheel).
# Each platform has its own loss profile (match rates, contextual blacklisting, etc.).
PLATFORMS = {
    "Semasio":   {"factor": 1.00, "note": "full motivational universe (source platform)"},
    "FreeWheel": {"factor": 0.70, "note": "~30% haircut (placeholder — real numbers TBD from FreeWheel)"},
}
DEFAULT_PLATFORM = "Semasio"


def _platform_factor(platform: str) -> float:
    return PLATFORMS.get(platform, PLATFORMS[DEFAULT_PLATFORM])["factor"]


@mcp.tool()
def add_to_campaign(campaign: str, creative_id: str, audience_id: str) -> str:
    """Group a creative-audience pair under a named campaign (e.g. 'Nike').

    A campaign is a set of (creative, audience) pairs. Call once per pair. This is
    what makes 'show me the Nike campaign' resolve later.
    """
    c = db.get_creative(creative_id)
    a = db.get_audience(audience_id)
    if not c:
        return json.dumps({"error": f"creative {creative_id} not found — score it first"})
    if not a:
        return json.dumps({"error": f"audience {audience_id} not found — create it first"})
    db.add_pair(campaign, creative_id, audience_id)
    db.log("add_to_campaign", f"{campaign}: {creative_id} x {audience_id}")
    return json.dumps({"campaign": campaign, "added": {"creative": c["name"], "audience": a["name"]}})


@mcp.tool()
def ingest_audience_universes(csv_text: str, platform: str, as_of: str) -> str:
    """Flow A (monthly refresh) — load per-motivation audience sizes for a platform.

    This is async reference data: audience sizes are loaded on a cadence (e.g. monthly),
    stored in the DB, and READ at request time — not computed live. Each platform
    (Semasio, FreeWheel, ...) has its own absolute sizes (FreeWheel's are already the
    smaller reachable numbers — no haircut applied on top of loaded data).

    csv_text columns: audience_id, motivation, size. platform + as_of (e.g. '2026-03')
    apply to the whole load. Motivations must be among the 16; others are flagged.
    """
    reader = csv.DictReader(io.StringIO(csv_text.strip()))
    loaded, flags = 0, []
    for row in reader:
        aud = (row.get("audience_id") or "").strip()
        mot = (row.get("motivation") or "").strip()
        if not aud or not mot:
            continue
        if mot not in MOTIVATIONS:
            flags.append(f"{aud}: '{mot}' not a valid motivation — skipped"); continue
        try:
            size = int(float(row.get("size", 0) or 0))
        except ValueError:
            flags.append(f"{aud}/{mot}: bad size — skipped"); continue
        db.upsert_audience_universe(aud, platform, mot, size, as_of)
        loaded += 1
    db.log("ingest_audience_universes", f"{loaded} rows, platform={platform}, as_of={as_of}")
    return json.dumps({"loaded": loaded, "platform": platform, "as_of": as_of, "flags": flags})


@mcp.tool()
def show_campaign(campaign: str, detail_for: str = "", platform: str = DEFAULT_PLATFORM) -> str:
    """User view of a campaign in the per-creative / per-motivation shape.

    Reads stored per-motivation audience sizes for the selected platform (loaded
    monthly via ingest_audience_universes). Table: one row per creative with its
    motivational profile, Audience Size (naive sum), Deduplicated Audience Size.
    Surfaces the data 'as_of' date so staleness is visible. If no data is loaded for
    a platform, falls back to clearly-labelled PLACEHOLDER numbers.

    platform: 'Semasio', 'FreeWheel', etc. detail_for: a creative_id for its spider +
    per-motivation breakdown.
    """
    view = db.campaign_view(campaign, platform, fallback_fn=_placeholder_universes)
    if view is None:
        return json.dumps({"error": f"no campaign named '{campaign}' — add pairs with add_to_campaign"})

    table, as_ofs, sources = [], set(), set()
    for r in view["creatives"]:
        if "error" in r:
            table.append({"creative": r.get("creative", r.get("creative_id")), "error": r["error"]})
            continue
        table.append({"creative": r["creative"], "creative_id": r["creative_id"],
                      "motivational_profile": r["motivational_profile"],
                      "audience_size": r["audience_size"],
                      "dedup_audience_size": r["dedup_audience_size"]})
        if r.get("as_of"): as_ofs.add(r["as_of"])
        sources.add(r["data_source"])
    out = {"campaign": campaign, "platform": platform,
           "as_of": sorted(as_ofs) or None, "data_source": sorted(sources),
           "available_platforms": list(PLATFORMS), "table": table, "creatives": len(table)}

    if detail_for:
        sel = next((r for r in view["creatives"] if r.get("creative_id") == detail_for and "error" not in r), None)
        if sel:
            out["detail"] = {
                "creative": sel["creative"], "audience": sel["audience"], "platform": platform,
                "as_of": sel.get("as_of"), "data_source": sel["data_source"],
                "spider_svg": polygon_svg(sel["creative_scores"], []),
                "per_motivation": sel["per_motivation"],
                "audience_size": sel["audience_size"],
                "dedup_audience_size": sel["dedup_audience_size"],
            }
        else:
            out["detail_error"] = f"{detail_for} not in campaign '{campaign}'"
    db.log("show_campaign", f"{campaign} platform={platform}")
    return json.dumps(out)




def _placeholder_universes(audience_id: str, total: int) -> dict:
    """FALLBACK ONLY — illustrative per-motivation sizes when no real data is loaded.
    Real sizes come from ingest_audience_universes (monthly). Clearly flagged in output
    as PLACEHOLDER so nobody mistakes these for loaded numbers."""
    total = total or 10_000_000
    out = {}
    for m in MOTIVATIONS:
        h = int.from_bytes(hashlib.sha256(f"{audience_id}|{m}".encode()).digest()[:4], "big")
        out[m] = int((0.08 + (h / 0xFFFFFFFF) * 0.20) * total)
    return out


@mcp.tool()
def programs_by_motivation(motivation: str, top_n: int = 20) -> str:
    """[INTERNAL ONLY] Rank stored TV programs by their score on one motivation.

    Browses the program-profile library — e.g. 'all shows high on Belonging'. Accepts
    natural words and maps them to the 16 (Family -> Belonging/Love, Achievement-y ->
    Achievement, etc.). INTERNAL because enumerating the program library by motivation
    exposes the proprietary content-scoring asset; do NOT expose externally.
    """
    SYN = {"family": "Belonging", "community": "Belonging", "togetherness": "Belonging",
           "romance": "Love", "relationships": "Love", "ambition": "Aspiration",
           "success": "Achievement", "fitness": "Fitness", "wealthy": "Wealth",
           "prestige": "Status", "self esteem": "Self-Esteem"}
    key = motivation.strip()
    if key not in MOTIVATIONS:
        mapped = SYN.get(key.lower())
        if not mapped:
            return json.dumps({"error": f"'{motivation}' is not one of the 16 motivations and has no mapping",
                               "valid": MOTIVATIONS})
        key = mapped
    progs = db.all_programs()
    ranked = sorted(progs, key=lambda p: p["scores"].get(key, 0), reverse=True)[:top_n]
    rows = [{"program": p["program"], "network": p["network"], "daypart": p["daypart"],
             "score_pct": round(p["scores"].get(key, 0) * 100), "reach": p["reach"]}
            for p in ranked]
    db.log("programs_by_motivation", f"{motivation}->{key}, top {len(rows)}")
    return json.dumps({"motivation": key, "requested_as": motivation, "programs": rows})


@mcp.tool()
def get_input_spec(spec: str = "ad_creative") -> str:
    """Return the declared input spec the database validates against.

    spec: currently 'ad_creative'. This is the single source of truth for what a
    valid creative looks like — fields, the 16 motivations, value ranges, rules.
    Claude reads this to map messy uploads toward the contract; validation uses it.
    """
    if spec in ("ad_creative", "creative"):
        return json.dumps(CREATIVE_INPUT_SPEC)
    return json.dumps({"error": "unknown spec", "available": ["ad_creative"]})


@mcp.tool()
def identify_database(filename: str, header_line: str) -> str:
    """1.1.2 — Identify which RMT database an inbound file belongs to."""
    name, reason = identify_database(filename, header_line)
    db.log("identify_database", f"{filename} -> {name} ({reason})")
    return json.dumps({"database": name, "reason": reason})


@mcp.tool()
def ingest(rows_json: str, known_genre: str = "") -> str:
    """1.1.1 — Clean + conform messy rows to the DB input spec (ENG translation).

    Returns conformed rows plus per-row flags. Rows flagged NEEDS REVIEW should be
    resolved by the agent before write_creative is called.
    """
    rows = json.loads(rows_json)
    out, all_flags = [], []
    for r in rows:
        res = T.conform_row(r, known_genre or None)
        out.append(res["row"])
        all_flags.extend(res["flags"])
    db.log("ingest", f"{len(rows)} rows, {len(all_flags)} flags")
    return json.dumps({"conformed": out, "flags": all_flags,
                       "needs_review": any("NEEDS REVIEW" in f for f in all_flags)})


@mcp.tool()
def score_creative(creative_ref: str, name: str = "", source_db: str = "US_CONTENT") -> str:
    """1.1.1 product front door — Score a RAW creative from a VAST tag URL or file.

    This is the user-facing flow: 'here's a creative, score it'. Accepts a VAST tag
    URL or an uploaded/local file path. Classifies the reference, scores it (dummy
    today; human Creative IE at v1.1; AI at v4.0), writes to the clean core, returns
    scores. asset_id is derived from the ref so re-scoring the same creative is stable.
    """
    kind, ref = classify_creative_ref(creative_ref)
    if kind == "unknown":
        db.log("score_creative", f"REJECTED unrecognized ref: {creative_ref[:80]}")
        return json.dumps({"error": "could not classify creative reference",
                           "hint": "expected a VAST tag URL (https://…vast…) or a file path",
                           "received": creative_ref[:120]})
    asset_id = "CR_" + hashlib.sha1(ref.encode()).hexdigest()[:10].upper()
    display = name or asset_id
    scores = db.score_and_store_creative(asset_id, display, ref=ref, source_db=source_db)
    db.log("score_creative", f"{kind}:{ref[:60]} -> {asset_id} ({display})")
    return json.dumps({"asset_id": asset_id, "name": display, "ref_kind": kind,
                       "scores": scores, "written": True})


def _notify(message: str) -> str:
    """Notification hook — POC stub. The call site is real; delivery is deferred.

    Today: logs the intent (no external send — single-person POC, no second human yet).
    Later (shared/remote server): swap this body for email / Slack / webhook to the
    scorer. Nothing else in the flow changes — same swap pattern as the engine/DB.
    """
    db.log("notify", message)
    return f"[notify-stub] would ping scorer: {message}"


def _resolve_creative(name_or_id: str):
    """Match a scorer-provided label to a queued/known creative, else mark as new."""
    # direct id hit (already in creative_scores)
    c = db.get_creative(name_or_id)
    if c:
        return {"creative_id": name_or_id, "name": c["name"], "match": "existing"}
    # a CR_ id that's queued but not yet scored -> use it as-is, look up its queue name
    if name_or_id.startswith("CR_"):
        qn = db.query("SELECT name FROM scoring_queue WHERE creative_id=? LIMIT 1", (name_or_id,))
        return {"creative_id": name_or_id,
                "name": qn[0]["name"] if qn else name_or_id, "match": "queued"}
    # queued by human name
    q = db.query("SELECT creative_id, name FROM scoring_queue WHERE lower(name)=lower(?) LIMIT 1",
                 (name_or_id,))
    if q:
        return {"creative_id": q[0]["creative_id"], "name": q[0]["name"], "match": "queued"}
    # else: new creative, derive an id from the label
    cid = "CR_" + hashlib.sha1(name_or_id.encode()).hexdigest()[:10].upper()
    return {"creative_id": cid, "name": name_or_id, "match": "new"}


SCORER_SETUP = """You are set up as a CREATIVE SCORER for ResonanceOS.

When I upload a scoring file (any layout — videos' filenames, motivations in rows or
columns, scores as 0-100 or 0-1, extra notes, anything):
0. If unsure of the target shape, call get_input_spec to read the ad_creative spec.
1. Read it and map every score column/label to exactly ONE of these 16 motivations:
   Transcendence, Self-Knowledge, Creativity, Love, Power, Good Life, Heroism, Wealth,
   Status, Fitness, Competency, Aspiration, Achievement, Belonging, Security, Self-Esteem.
2. Normalize all scores to 0-100. Pull each creative's name from its filename.
3. ALWAYS call preview_scored_creatives first and show me a clean summary of what you
   understood (per creative: scores, whether it matched the queue or is new, any flags).
   Never write to the database before I confirm.
4. If any label can't map to one of the 16 motivations, tell me and ask — do not guess.
5. Only after I say yes, call commit_scored_creatives.
Also: at the start of each session, call view_scoring_queue and tell me what's pending."""

REQUESTER_SETUP = """You are set up as a REQUESTER for ResonanceOS.

When I paste creative links or filenames and ask to score them:
1. Call enqueue_for_scoring once per creative, with requested_by set to my name.
2. Confirm back which creatives were queued and how many are now pending.
Keep it to that — I'm requesting scoring, not doing it."""

USER_SETUP = """You are set up as a REPORT USER for ResonanceOS.

When I ask about a campaign or creatives:
1. Use show_campaign to get the table (creatives, motivational profiles, audience size,
   deduplicated audience size). Pass the platform I name (e.g. Semasio, FreeWheel);
   default Semasio. Use detail_for=<creative_id> for one creative's spider + breakdown.
2. Use generate_report for the AD_MOT Excel (Programs / Network / Daypart).

You may freely discuss, rank, compare, explain, and summarize whatever the tools
return — that's reading data I already have.

CRITICAL — never invent or recalculate numbers yourself. Every audience size, dedup
figure, match, or universe number must come from a tool call, not your own arithmetic.
If I ask for a different platform, threshold, audience, or any figure not in the last
result, CALL THE TOOL AGAIN to get it. If no tool can produce it, say so plainly —
do not estimate. Also surface the data's 'as_of' date so I know how current it is."""


@mcp.tool()
def setup_instructions(role: str = "scorer") -> str:
    """Return the project-setup text to paste into your Claude project settings.

    Ask 'get me set up right' and pass your role: 'scorer', 'requester', or 'user'.
    Copy the returned text into Settings → (your project) → custom instructions. The
    server can't write it for you, but this is the exact text to paste — one time.
    """
    texts = {"scorer": SCORER_SETUP, "requester": REQUESTER_SETUP, "user": USER_SETUP}
    role = role.lower().strip()
    if role not in texts:
        return json.dumps({"error": "role must be scorer, requester, or user",
                           "available": list(texts)})
    return json.dumps({
        "role": role,
        "paste_into": "Settings → your Claude project → custom instructions",
        "instruction_text": texts[role],
        "note": "Paste the instruction_text once. After that, just talk normally — "
                "e.g. a scorer only needs to say 'these are scored, put them in the database'."
    })


@mcp.tool()
def preview_scored_creatives(scores_json: str) -> str:
    """Step 1 of scorer intake — validate + preview, WRITE NOTHING.

    Claude reads the scorer's (free-form) Excel, maps it to the 16-motivation spec,
    and calls this with clean structured data:
      [{"creative": "<name or id>", "scores": {"Status": 48, "Wealth": 72, ...}}, ...]
    Returns 'here is what I understood' — per creative: resolved id, match type
    (existing/queued/new), the parsed 0-100 scores, and any flags — for the scorer
    to confirm before commit. Rejects unknown motivations / out-of-range values.
    """
    try:
        items = json.loads(scores_json)
    except Exception:
        return json.dumps({"error": "scores_json must be JSON: [{creative, scores{...}}, ...]"})
    preview, flags = [], []
    for it in items:
        label = str(it.get("creative", "")).strip()
        raw = it.get("scores", {}) or {}
        if not label:
            flags.append("an item has no creative name/id — skipped"); continue
        item_problems = validate_creative(it)
        res = _resolve_creative(label)
        row = {"creative": label, **res,
               "scores_pct": {m: raw.get(m, 0) for m in MOTIVATIONS if raw.get(m, 0)}}
        if item_problems:
            row["problems"] = item_problems
            flags.extend(item_problems)
        if res["match"] == "new":
            flags.append(f"{label}: no queued match — will be created as NEW creative")
        preview.append(row)
    db.log("preview_scored_creatives", f"{len(preview)} creatives, {len(flags)} flags")
    return json.dumps({"understood": preview, "flags": flags,
                       "next": "If correct, call commit_scored_creatives with the same scores_json."})


@mcp.tool()
def commit_scored_creatives(scores_json: str, scored_by: str = "scorer") -> str:
    """Step 2 of scorer intake — write the previewed scores to the DB.

    Same scores_json as preview. Resolves each creative (match queued or create new),
    stores its graded profile, and marks any matching queue item scored. Use only after
    the scorer has confirmed the preview.
    """
    try:
        items = json.loads(scores_json)
    except Exception:
        return json.dumps({"error": "scores_json must be JSON: [{creative, scores{...}}, ...]"})

    # HARD GATE via the declared spec. No partial writes, no silent skips.
    problems = []
    for it in items:
        problems.extend(validate_creative(it))
    if problems:
        return json.dumps({"error": "commit blocked — resolve these in the Excel and re-preview, "
                           "then commit again", "problems": problems, "written": 0})

    written = []
    for it in items:
        label = str(it.get("creative", "")).strip()
        raw = it.get("scores", {}) or {}
        res = _resolve_creative(label)
        scores = {m: round(float(raw.get(m, 0)) / 100.0, 3) for m in MOTIVATIONS}
        db.store_graded_creative(res["creative_id"], res["name"], scores)
        q = db.query("SELECT queue_id FROM scoring_queue WHERE creative_id=? AND status='pending' LIMIT 1",
                     (res["creative_id"],))
        if q:
            db.mark_scored(q[0]["queue_id"], scored_by)
        written.append({"creative_id": res["creative_id"], "name": res["name"], "match": res["match"]})
    db.log("commit_scored_creatives", f"{len(written)} by {scored_by}: {[w['creative_id'] for w in written]}")
    return json.dumps({"written": written, "count": len(written), "scored_by": scored_by})


@mcp.tool()
def enqueue_for_scoring(filename: str, video_url: str = "", requested_by: str = "requester") -> str:
    """Add a creative to the scoring queue. Mines metadata from the filename.

    requested_by: who is requesting the scoring (e.g. the requester's name).
    The queue item carries identity + mined metadata so the scorer never transcribes
    an ID (anti-misattribution). Fires the notification hook to the scorer (stub today).
    """
    meta = mine_filename(filename)
    creative_id = "CR_" + hashlib.sha1(filename.encode()).hexdigest()[:10].upper()
    queue_id = "Q_" + hashlib.sha1((filename + str(time.time())).encode()).hexdigest()[:8].upper()
    name = meta.get("product", filename)
    db.enqueue(queue_id, creative_id, name, meta, video_url, requested_by)
    db.log("enqueue_for_scoring", f"{queue_id} -> {creative_id} ({name}) by {requested_by}")
    pending = len(db.pending_queue())
    notice = _notify(f"{requested_by} requested scoring of '{name}' ({pending} pending in queue)")
    return json.dumps({"queue_id": queue_id, "creative_id": creative_id, "requested_by": requested_by,
                       "metadata": meta, "status": "pending", "notification": notice})


@mcp.tool()
def view_scoring_queue() -> str:
    """Show all creatives pending scoring (Skip's worklist)."""
    items = db.pending_queue()
    for it in items:
        it["metadata"] = json.loads(it.pop("metadata_json"))
    db.log("view_scoring_queue", f"{len(items)} pending")
    return json.dumps({"pending": items})


@mcp.tool()
def open_scoring_card(queue_id: str) -> str:
    """Open one queue item: parsed description, video link, and the 16 motivations to check."""
    item = db.get_queue_item(queue_id)
    if not item:
        return json.dumps({"error": f"queue item {queue_id} not found"})
    meta = json.loads(item["metadata_json"])
    db.log("open_scoring_card", queue_id)
    return json.dumps({
        "queue_id": queue_id,
        "creative": item["name"],
        "requested_by": item.get("requested_by", "unknown"),
        "description": meta,
        "video_url": item["video_url"],
        "motivations_to_check": MOTIVATIONS,
        "note": "Watch the video, then submit_scores with the queue_id and the "
                "motivations that apply. Identity is bound to this card — do not pass a creative id.",
    })


@mcp.tool()
def submit_scores(queue_id: str, scores_json: str, scored_by: str = "skip") -> str:
    """Submit GRADED scores (0-100% per motivation) for the creative bound to this card.

    scores_json: JSON object {motivation: 0-100}. Motivations omitted are treated as 0.
    Identity comes from the queue_id, NOT the caller — prevents misattribution.
    Unknown motivation names and out-of-range values are rejected, not guessed.
    Graded scores are what give the polygon its shape; the 20%+ activation
    threshold is applied later, at view/decision time, NOT stored here.
    """
    item = db.get_queue_item(queue_id)
    if not item:
        return json.dumps({"error": f"queue item {queue_id} not found"})
    if item["status"] == "scored":
        return json.dumps({"error": f"{queue_id} already scored by {item['scored_by']}"})

    try:
        raw = json.loads(scores_json)
    except Exception:
        return json.dumps({"error": "scores_json must be a JSON object {motivation: 0-100}"})
    unknown = set(raw) - set(MOTIVATIONS)
    if unknown:
        return json.dumps({"error": "unknown motivations", "unknown": sorted(unknown),
                           "valid": MOTIVATIONS})
    bad = {k: v for k, v in raw.items() if not (isinstance(v, (int, float)) and 0 <= v <= 100)}
    if bad:
        return json.dumps({"error": "scores must be numbers 0-100", "out_of_range": bad})

    # store as 0-1 ratios in the clean core
    scores = {m: round(float(raw.get(m, 0)) / 100.0, 3) for m in MOTIVATIONS}
    creative_id = item["creative_id"]
    db.store_graded_creative(creative_id, item["name"], scores)
    db.mark_scored(queue_id, scored_by)
    nonzero = {m: raw[m] for m in raw if raw.get(m, 0) > 0}
    db.log("submit_scores", f"{queue_id} -> {creative_id} by {scored_by}: {nonzero}")
    return json.dumps({"creative_id": creative_id, "name": item["name"],
                       "scores_pct": {m: round(scores[m]*100) for m in MOTIVATIONS},
                       "scored_by": scored_by, "written": True})


@mcp.tool()
def ingest_program_profiles(csv_text: str) -> str:
    """Stage-1 RMT DB load — upload pre-scored program profiles as CSV.

    Expected columns: program, network, daypart, reach, then one column per
    motivation (0-100). Header row required. Motivation columns are matched by
    name to the 16 RMT motivations; missing ones default to 0. This is the
    'update the RMT database' capability — programs are the pre-scored reference
    profiles the reports compute against.
    """
    reader = csv.DictReader(io.StringIO(csv_text.strip()))
    loaded, flags = 0, []
    for row in reader:
        prog = (row.get("program") or "").strip()
        if not prog:
            continue
        scores = {}
        for m in MOTIVATIONS:
            v = row.get(m, row.get(m.lower(), ""))
            try:
                scores[m] = round(float(v) / 100.0, 3) if str(v).strip() != "" else 0.0
            except ValueError:
                scores[m] = 0.0
                flags.append(f"{prog}: bad value for {m} -> 0")
        try:
            reach = int(float(row.get("reach", 0) or 0))
        except ValueError:
            reach = 0
        db.upsert_program(prog, (row.get("network") or "").strip(),
                          (row.get("daypart") or "").strip(), scores, reach)
        loaded += 1
    db.log("ingest_program_profiles", f"{loaded} programs, {len(flags)} flags")
    return json.dumps({"loaded": loaded, "flags": flags,
                       "total_in_db": len(db.all_programs())})


@mcp.tool()
def view_program_profiles() -> str:
    """Show the pre-scored program profiles currently in the RMT DB."""
    progs = db.all_programs()
    summary = [{"program": p["program"], "network": p["network"],
                "daypart": p["daypart"], "reach": p["reach"],
                "top_motivations": sorted(p["scores"], key=p["scores"].get, reverse=True)[:3]}
               for p in progs]
    return json.dumps({"count": len(progs), "programs": summary})


@mcp.tool()
def generate_report(creative_id: str, report_types: str = "all") -> str:
    """1.1.3 — Generate AD_MOT standard reports as a downloadable Excel workbook.

    Computes the creative's resonance against the STORED program profiles in the
    RMT DB (not invented numbers). Requires program profiles to be loaded first
    via ingest_program_profiles. Returns the workbook path.
    """
    c = db.get_creative(creative_id)
    if not c:
        return json.dumps({"error": f"creative {creative_id} not found — score it first"})
    programs = db.all_programs()
    if not programs:
        return json.dumps({"error": "no program profiles in DB — load them first via ingest_program_profiles"})
    creative_scores = json.loads(c["scores_json"])
    meta = {}
    q = db.query("SELECT metadata_json FROM scoring_queue WHERE creative_id=? LIMIT 1", (creative_id,))
    if q:
        meta = json.loads(q[0]["metadata_json"])
    path = excel_report.build_workbook(creative_id, c["name"], creative_scores, programs, meta)
    db.log("generate_report", f"{creative_id} vs {len(programs)} programs -> {path}")
    return json.dumps({"creative_id": creative_id, "name": c["name"],
                       "programs_evaluated": len(programs), "workbook_path": path,
                       "tabs": ["Summary", "By Program", "By Network", "By Network + Daypart"]})


@mcp.tool()
def write_creative(asset_id: str, name: str, source_db: str = "US_CONTENT") -> str:
    """1.1.1b — Score a pre-identified creative by id and write it (internal/data-load path)."""
    scores = db.score_and_store_creative(asset_id, name, source_db=source_db)
    db.log("write_creative", f"{asset_id} ({name})")
    return json.dumps({"asset_id": asset_id, "scores": scores, "written": True})


@mcp.tool()
def write_audience(audience_id: str, name: str, universe_size: int) -> str:
    """Score + store an audience profile (section 2)."""
    scores = db.score_and_store_audience(audience_id, name, universe_size)
    db.log("write_audience", f"{audience_id} ({name}), universe={universe_size}")
    return json.dumps({"audience_id": audience_id, "scores": scores,
                       "universe_size": universe_size, "written": True})


@mcp.tool()
def score_match(audience_id: str, creative_ids: str) -> str:
    """1.1.9 + 1.1.10 — Resonance match: rank creatives vs an audience, return polygon.

    creative_ids: comma-separated asset_ids already written to the DB.
    Returns alignment summary per creative AND an inline SVG polygon (the MCP's
    hero output format).
    """
    ids = [c.strip() for c in creative_ids.split(",") if c.strip()]
    m = db.match(audience_id, ids)        # DB owns the alignment computation
    if m is None:
        return json.dumps({"error": f"audience {audience_id} not found"})

    # engine only FORMATS output: build the polygon from DB results
    overlays = [(r["name"], r["scores"], r) for r in m["results"] if "scores" in r]
    svg = polygon_svg(m["audience_scores"], overlays)
    clean = [{k: v for k, v in r.items() if k != "scores"} for r in m["results"]]
    db.log("score_match", f"audience={audience_id}, creatives={ids}")
    return json.dumps({"audience": m["audience"], "results": clean, "polygon_svg": svg})


@mcp.tool()
def query_db(sql: str) -> str:
    """1.1.7 (internal NLQ substrate) — run a read query against the clean core."""
    if not sql.strip().lower().startswith("select"):
        return json.dumps({"error": "read-only: SELECT statements only"})
    db.log("query_db", sql[:120])
    return json.dumps({"rows": db.query(sql)})


@mcp.tool()
def audit_tail(n: int = 20) -> str:
    """1.1.6 — Return the last n audit-log entries."""
    rows = db.query("SELECT ts, operation, detail FROM audit_log ORDER BY id DESC LIMIT ?", (n,))
    return json.dumps({"entries": rows})


def main():
    transport = os.environ.get("RESONANCE_TRANSPORT", "stdio").lower()
    if transport == "http":
        # shared server: both requester and scorer connect to one URL -> one queue.
        # managed hosts (Render/Railway) inject PORT and expect binding on 0.0.0.0
        host = os.environ.get("RESONANCE_HOST", "0.0.0.0")
        port = int(os.environ.get("PORT", os.environ.get("RESONANCE_PORT", "8848")))
        mcp.settings.host = host
        mcp.settings.port = port
        mcp.run(transport="streamable-http")
    else:
        mcp.run()  # stdio (local POC, single machine)


if __name__ == "__main__":
    main()

